"""
Generate Professional UT CML Templates for NDE Companies
Creates polished, field-ready Excel templates with formulas, validation, and formatting.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, Protection,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from copy import copy

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Color Palette - Professional Industrial Theme
COLORS = {
    'primary_dark': '1F4E79',      # Dark blue - main headers
    'primary_light': 'D6E3F8',     # Light blue - section backgrounds
    'accent': '2E75B6',            # Medium blue - subheaders
    'header_bg': 'BDD7EE',         # Light blue - table headers
    'input_bg': 'FFFFFF',          # White - input cells
    'calculated_bg': 'FFF2CC',     # Light yellow - formula cells
    'warning_bg': 'FCE4D6',        # Light orange - warning
    'critical_bg': 'F8CBAD',       # Orange-red - critical
    'success_bg': 'E2EFDA',        # Light green - acceptable
    'border': '7F7F7F',            # Gray - borders
    'text_dark': '1F1F1F',         # Near black - main text
    'text_light': 'FFFFFF',        # White - on dark backgrounds
}

# Border Styles
thin_border = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='thin', color=COLORS['border']),
    bottom=Side(style='thin', color=COLORS['border'])
)

medium_border = Border(
    left=Side(style='medium', color=COLORS['border']),
    right=Side(style='medium', color=COLORS['border']),
    top=Side(style='medium', color=COLORS['border']),
    bottom=Side(style='medium', color=COLORS['border'])
)

# Font Styles
font_title = Font(name='Calibri', size=14, bold=True, color=COLORS['text_light'])
font_section = Font(name='Calibri', size=11, bold=True, color=COLORS['text_dark'])
font_header = Font(name='Calibri', size=10, bold=True, color=COLORS['text_dark'])
font_label = Font(name='Calibri', size=10, bold=True, color=COLORS['text_dark'])
font_normal = Font(name='Calibri', size=10, color=COLORS['text_dark'])
font_small = Font(name='Calibri', size=9, color=COLORS['text_dark'])
font_company = Font(name='Calibri', size=16, bold=True, color=COLORS['primary_dark'])

# Fill Styles
fill_title = PatternFill(start_color=COLORS['primary_dark'], end_color=COLORS['primary_dark'], fill_type='solid')
fill_header = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
fill_section = PatternFill(start_color=COLORS['primary_light'], end_color=COLORS['primary_light'], fill_type='solid')
fill_input = PatternFill(start_color=COLORS['input_bg'], end_color=COLORS['input_bg'], fill_type='solid')
fill_calculated = PatternFill(start_color=COLORS['calculated_bg'], end_color=COLORS['calculated_bg'], fill_type='solid')
fill_critical = PatternFill(start_color=COLORS['critical_bg'], end_color=COLORS['critical_bg'], fill_type='solid')
fill_warning = PatternFill(start_color=COLORS['warning_bg'], end_color=COLORS['warning_bg'], fill_type='solid')
fill_success = PatternFill(start_color=COLORS['success_bg'], end_color=COLORS['success_bg'], fill_type='solid')

# Alignment
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def set_column_widths(ws, widths):
    """Set column widths from a dictionary {column_letter: width}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_height(ws, row, height):
    """Set row height"""
    ws.row_dimensions[row].height = height


def style_cell(cell, font=None, fill=None, border=None, alignment=None):
    """Apply styles to a cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment


def create_header_block(ws, title, start_row=1):
    """
    Create standardized company header block
    Returns the next available row after the header
    """
    # Row 1: Company Logo Area (placeholder) + Title
    ws.merge_cells(f'A{start_row}:B{start_row + 2}')
    logo_cell = ws.cell(row=start_row, column=1)
    logo_cell.value = "[COMPANY LOGO]"
    style_cell(logo_cell, font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    ws.merge_cells(f'C{start_row}:G{start_row}')
    company_cell = ws.cell(row=start_row, column=3)
    company_cell.value = "YOUR COMPANY NAME"
    style_cell(company_cell, font=font_company, border=thin_border, alignment=align_center)

    # Row 2: Document Title
    ws.merge_cells(f'C{start_row + 1}:G{start_row + 1}')
    title_cell = ws.cell(row=start_row + 1, column=3)
    title_cell.value = title
    style_cell(title_cell, font=font_title, fill=fill_title, border=thin_border, alignment=align_center)

    # Row 3: Document Info
    ws.merge_cells(f'C{start_row + 2}:D{start_row + 2}')
    doc_num_label = ws.cell(row=start_row + 2, column=3)
    doc_num_label.value = "Doc No:"
    style_cell(doc_num_label, font=font_label, fill=fill_section, border=thin_border, alignment=align_left)

    ws.merge_cells(f'E{start_row + 2}:F{start_row + 2}')
    rev_label = ws.cell(row=start_row + 2, column=5)
    rev_label.value = "Rev:"
    style_cell(rev_label, font=font_label, fill=fill_section, border=thin_border, alignment=align_left)

    page_cell = ws.cell(row=start_row + 2, column=7)
    page_cell.value = "Page:"
    style_cell(page_cell, font=font_label, fill=fill_section, border=thin_border, alignment=align_left)

    # Row 4: Inspection Info
    row4 = start_row + 3
    labels = ['Inspector:', 'Date:', 'Equipment ID:', 'Client:']
    for i, label in enumerate(labels):
        col = i * 2 + 1
        ws.cell(row=row4, column=col).value = label
        style_cell(ws.cell(row=row4, column=col), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        style_cell(ws.cell(row=row4, column=col + 1), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Set row heights
    for r in range(start_row, start_row + 4):
        set_row_height(ws, r, 22)

    return start_row + 5  # Return next available row


def create_signature_block(ws, start_row, start_col=1, num_cols=8):
    """Create signature/approval block at bottom of sheet"""
    ws.merge_cells(f'{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + num_cols - 1)}{start_row}')
    sig_header = ws.cell(row=start_row, column=start_col)
    sig_header.value = "APPROVAL / SIGN-OFF"
    style_cell(sig_header, font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    # Signature rows
    sig_labels = ['Inspected By:', 'Reviewed By:', 'Approved By:']
    for i, label in enumerate(sig_labels):
        row = start_row + 1 + i
        ws.cell(row=row, column=start_col).value = label
        style_cell(ws.cell(row=row, column=start_col), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)

        ws.merge_cells(f'{get_column_letter(start_col + 1)}{row}:{get_column_letter(start_col + 2)}{row}')
        style_cell(ws.cell(row=row, column=start_col + 1), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

        ws.cell(row=row, column=start_col + 3).value = "Signature:"
        style_cell(ws.cell(row=row, column=start_col + 3), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)

        ws.merge_cells(f'{get_column_letter(start_col + 4)}{row}:{get_column_letter(start_col + 5)}{row}')
        style_cell(ws.cell(row=row, column=start_col + 4), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

        ws.cell(row=row, column=start_col + 6).value = "Date:"
        style_cell(ws.cell(row=row, column=start_col + 6), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        style_cell(ws.cell(row=row, column=start_col + 7), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

        set_row_height(ws, row, 25)

    return start_row + 4


def setup_print_settings(ws, orientation='portrait', fit_to_page=True):
    """Configure print settings for professional output"""
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = fit_to_page
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # As many pages as needed vertically

    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75,
        header=0.3, footer=0.3
    )

    # Add footer with page numbers
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.right.text = "&D"


def add_conditional_formatting_thickness(ws, meas_col, tmin_col, start_row, end_row):
    """Add conditional formatting: red if measured < Tmin"""
    meas_letter = get_column_letter(meas_col)
    tmin_letter = get_column_letter(tmin_col)

    # Red fill if measured thickness is less than minimum
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    red_font = Font(color='9C0006', bold=True)

    for row in range(start_row, end_row + 1):
        formula = f'AND({meas_letter}{row}<>"", {tmin_letter}{row}<>"", {meas_letter}{row}<{tmin_letter}{row})'
        ws.conditional_formatting.add(
            f'{meas_letter}{row}',
            FormulaRule(formula=[formula], fill=red_fill, font=red_font)
        )


# =============================================================================
# TEMPLATE CREATORS
# =============================================================================

def create_piping_circuit_sheet(wb):
    """1. Piping Circuit Thickness Sheet (API 570 / 574)"""
    ws = wb.create_sheet("01_PipingCircuit")

    # Column widths
    set_column_widths(ws, {
        'A': 12, 'B': 30, 'C': 12, 'D': 12, 'E': 12, 'F': 10, 'G': 10, 'H': 25
    })

    # Header block
    next_row = create_header_block(ws, "PIPING CIRCUIT THICKNESS SHEET (API 570/574)")

    # Equipment Information Section
    info_row = next_row
    ws.merge_cells(f'A{info_row}:H{info_row}')
    ws.cell(row=info_row, column=1).value = "CIRCUIT INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)
    set_row_height(ws, info_row, 20)

    # Info fields row 1
    info_row += 1
    info_fields = [
        ('Line No:', 'A', 'B'),
        ('Circuit ID:', 'C', 'D'),
        ('P&ID Ref:', 'E', 'F'),
        ('Service:', 'G', 'H'),
    ]
    for label, label_col, value_col in info_fields:
        ws[f'{label_col}{info_row}'] = label
        style_cell(ws[f'{label_col}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        style_cell(ws[f'{value_col}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
    set_row_height(ws, info_row, 22)

    # Info fields row 2
    info_row += 1
    info_fields2 = [
        ('NPS:', 'A', 'B'),
        ('Schedule:', 'C', 'D'),
        ('Material:', 'E', 'F'),
        ('Nom. t (mm):', 'G', 'H'),
    ]
    for label, label_col, value_col in info_fields2:
        ws[f'{label_col}{info_row}'] = label
        style_cell(ws[f'{label_col}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        style_cell(ws[f'{value_col}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
    set_row_height(ws, info_row, 22)

    # Blank row
    info_row += 1

    # Data table header
    table_header_row = info_row + 1
    headers = ['CML', 'Location Description', 'Nom t\n(mm)', 'Meas t\n(mm)', 'Tmin\n(mm)', 'CR\n(mm/yr)', 'RL\n(yrs)', 'Notes / Observations']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_header_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)
    set_row_height(ws, table_header_row, 30)

    # Data rows (20 rows for data entry)
    data_start_row = table_header_row + 1
    data_rows = 20
    for i in range(data_rows):
        row = data_start_row + i
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            # CR and RL columns get formula placeholder styling
            if col in [6, 7]:
                style_cell(cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
            else:
                style_cell(cell, font=font_normal, fill=fill_input, border=thin_border, alignment=align_center if col != 2 and col != 8 else align_left)

        # Add MIN formula for column 4 if multiple readings existed (placeholder)
        # Add CR formula: CR = (Last t - Curr t) / Years - simplified placeholder
        set_row_height(ws, row, 22)

    # Add conditional formatting for measured < Tmin
    add_conditional_formatting_thickness(ws, 4, 5, data_start_row, data_start_row + data_rows - 1)

    # Signature block
    sig_row = data_start_row + data_rows + 1
    create_signature_block(ws, sig_row)

    # Freeze panes
    ws.freeze_panes = f'A{table_header_row + 1}'

    # Print settings
    setup_print_settings(ws, orientation='landscape')
    ws.print_title_rows = f'1:{table_header_row}'

    return ws


def create_orientation_grid_sheet(wb):
    """2. Orientation-Based CML Grid"""
    ws = wb.create_sheet("02_OrientationGrid")

    set_column_widths(ws, {'A': 18, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 20})

    next_row = create_header_block(ws, "ORIENTATION-BASED CML GRID")

    # Equipment info
    info_row = next_row
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws.cell(row=info_row, column=1).value = "LOCATION INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields = [('Line No:', 'A', 'B'), ('CML ID:', 'C', 'D'), ('Location:', 'E', 'G')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc != 'G':
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Orientation diagram placeholder
    info_row += 2
    ws.merge_cells(f'A{info_row}:C{info_row + 4}')
    diagram_cell = ws.cell(row=info_row, column=1)
    diagram_cell.value = "ORIENTATION DIAGRAM\n\n     12:00 (Top)\n9:00  O  3:00\n     6:00 (Bottom)"
    style_cell(diagram_cell, font=font_normal, fill=fill_section, border=thin_border, alignment=align_center)

    # Orientation readings table
    ws.merge_cells(f'D{info_row}:G{info_row}')
    ws.cell(row=info_row, column=4).value = "THICKNESS READINGS (mm)"
    style_cell(ws.cell(row=info_row, column=4), font=font_section, fill=fill_header, border=thin_border, alignment=align_center)

    orientations = ["12 o'clock (Top)", "3 o'clock", "6 o'clock (Bottom)", "9 o'clock"]
    for i, orient in enumerate(orientations):
        row = info_row + 1 + i
        ws.cell(row=row, column=4).value = orient
        style_cell(ws.cell(row=row, column=4), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        ws.merge_cells(f'E{row}:G{row}')
        style_cell(ws.cell(row=row, column=5), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        set_row_height(ws, row, 22)

    # Summary section
    summary_row = info_row + 6
    ws.merge_cells(f'A{summary_row}:G{summary_row}')
    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    style_cell(ws.cell(row=summary_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    summary_fields = [
        ('Minimum Reading (mm):', 'Calculated'),
        ('Nominal Thickness (mm):', ''),
        ('Tmin (mm):', ''),
        ('Status:', 'ACCEPTABLE / MONITOR / REPLACE'),
    ]
    for i, (label, hint) in enumerate(summary_fields):
        row = summary_row + 1 + i
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1).value = label
        style_cell(ws.cell(row=row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        ws.merge_cells(f'D{row}:G{row}')
        cell = ws.cell(row=row, column=4)
        if i == 0:
            # MIN formula placeholder
            style_cell(cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        else:
            style_cell(cell, font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        set_row_height(ws, row, 22)

    # Add data validation for Status field
    status_dv = DataValidation(type="list", formula1='"ACCEPTABLE,MONITOR,REPLACE"', allow_blank=True)
    status_dv.error = "Please select from the list"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(ws.cell(row=summary_row + 4, column=4))

    # Comments section
    comments_row = summary_row + 6
    ws.merge_cells(f'A{comments_row}:G{comments_row}')
    ws.cell(row=comments_row, column=1).value = "COMMENTS / OBSERVATIONS"
    style_cell(ws.cell(row=comments_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    ws.merge_cells(f'A{comments_row + 1}:G{comments_row + 4}')
    style_cell(ws.cell(row=comments_row + 1, column=1), font=font_normal, fill=fill_input, border=thin_border, alignment=Alignment(horizontal='left', vertical='top', wrap_text=True))

    # Signature
    sig_row = comments_row + 6
    create_signature_block(ws, sig_row, num_cols=7)

    setup_print_settings(ws, orientation='portrait')

    return ws


def create_pipe_shoe_cui_sheet(wb):
    """3. Pipe Shoe / CUI Thickness Sheet"""
    ws = wb.create_sheet("03_PipeShoe_CUI")

    set_column_widths(ws, {'A': 15, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 20})

    next_row = create_header_block(ws, "PIPE SHOE / CUI THICKNESS SHEET")

    # Info section
    info_row = next_row
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws.cell(row=info_row, column=1).value = "LOCATION INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields1 = [('Line No:', 'A', 'B'), ('Pipe Shoe ID:', 'C', 'D'), ('Location:', 'E', 'G')]
    for label, lc, vc in fields1:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    info_row += 1
    ws.cell(row=info_row, column=1).value = "Support Type:"
    style_cell(ws.cell(row=info_row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
    ws.merge_cells(f'B{info_row}:C{info_row}')
    style_cell(ws.cell(row=info_row, column=2), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    ws.cell(row=info_row, column=4).value = "Insulation Removed:"
    style_cell(ws.cell(row=info_row, column=4), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
    style_cell(ws.cell(row=info_row, column=5), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)

    # Add Y/N validation
    yn_dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    yn_dv.add(ws.cell(row=info_row, column=5))

    ws.cell(row=info_row, column=6).value = "CUI Observed:"
    style_cell(ws.cell(row=info_row, column=6), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
    style_cell(ws.cell(row=info_row, column=7), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
    yn_dv.add(ws.cell(row=info_row, column=7))

    # Thickness readings table
    table_row = info_row + 2
    ws.merge_cells(f'A{table_row}:G{table_row}')
    ws.cell(row=table_row, column=1).value = "THICKNESS READINGS (mm)"
    style_cell(ws.cell(row=table_row, column=1), font=font_section, fill=fill_header, border=thin_border, alignment=align_center)

    table_row += 1
    headers = ['Position', 'Reading 1', 'Reading 2', 'Reading 3', 'Average', 'Tmin', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)

    positions = ['Top', 'Bottom', 'Side A (Left)', 'Side B (Right)', 'At Support']
    table_row += 1
    for i, pos in enumerate(positions):
        row = table_row + i
        ws.cell(row=row, column=1).value = pos
        style_cell(ws.cell(row=row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        for col in range(2, 5):  # Reading columns
            style_cell(ws.cell(row=row, column=col), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        # Average formula
        avg_cell = ws.cell(row=row, column=5)
        avg_cell.value = f'=IF(COUNT(B{row}:D{row})>0,AVERAGE(B{row}:D{row}),"")'
        style_cell(avg_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        avg_cell.number_format = '0.00'
        # Tmin
        style_cell(ws.cell(row=row, column=6), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        # Status formula
        status_cell = ws.cell(row=row, column=7)
        status_cell.value = f'=IF(OR(E{row}="",F{row}=""),"",IF(E{row}>=F{row},"OK","BELOW Tmin"))'
        style_cell(status_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        set_row_height(ws, row, 22)

    # Summary
    summary_row = table_row + len(positions) + 1
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Minimum Thickness Found (mm):"
    style_cell(ws.cell(row=summary_row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
    min_cell = ws.cell(row=summary_row, column=4)
    min_cell.value = f'=IF(COUNT(E{table_row}:E{table_row + len(positions) - 1})>0,MIN(E{table_row}:E{table_row + len(positions) - 1}),"")'
    style_cell(min_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
    min_cell.number_format = '0.00'

    # Observations
    obs_row = summary_row + 2
    ws.merge_cells(f'A{obs_row}:G{obs_row}')
    ws.cell(row=obs_row, column=1).value = "OBSERVATIONS / DEFECTS NOTED"
    style_cell(ws.cell(row=obs_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    ws.merge_cells(f'A{obs_row + 1}:G{obs_row + 5}')
    style_cell(ws.cell(row=obs_row + 1, column=1), font=font_normal, fill=fill_input, border=thin_border, alignment=Alignment(horizontal='left', vertical='top', wrap_text=True))

    sig_row = obs_row + 7
    create_signature_block(ws, sig_row, num_cols=7)

    setup_print_settings(ws, orientation='portrait')

    return ws


def create_isometric_legend_sheet(wb):
    """4. Isometric + CML Legend"""
    ws = wb.create_sheet("04_Isometric_Legend")

    set_column_widths(ws, {'A': 12, 'B': 18, 'C': 15, 'D': 15, 'E': 15, 'F': 15})

    next_row = create_header_block(ws, "ISOMETRIC SKETCH + CML LEGEND")

    # Info
    info_row = next_row
    fields = [('Drawing/ISO Ref:', 'A', 'C'), ('System:', 'D', 'F')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        end_col = chr(ord(vc) - 1) if vc != 'C' else 'C'
        if lc == 'A':
            ws.merge_cells(f'B{info_row}:C{info_row}')
            style_cell(ws[f'B{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            ws.merge_cells(f'E{info_row}:F{info_row}')
            style_cell(ws[f'E{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Sketch area
    sketch_row = info_row + 2
    ws.merge_cells(f'A{sketch_row}:F{sketch_row}')
    ws.cell(row=sketch_row, column=1).value = "ISOMETRIC SKETCH AREA"
    style_cell(ws.cell(row=sketch_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    sketch_row += 1
    ws.merge_cells(f'A{sketch_row}:F{sketch_row + 14}')
    sketch_cell = ws.cell(row=sketch_row, column=1)
    sketch_cell.value = "(Inspector draws piping run and marks CML locations)\n\nUse arrows to indicate flow direction\nMark CML numbers at measurement locations\nIndicate supports, fittings, and key features"
    style_cell(sketch_cell, font=font_normal, fill=fill_input, border=medium_border, alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))
    for r in range(sketch_row, sketch_row + 15):
        set_row_height(ws, r, 22)

    # CML Legend table
    legend_row = sketch_row + 16
    ws.merge_cells(f'A{legend_row}:F{legend_row}')
    ws.cell(row=legend_row, column=1).value = "CML LEGEND"
    style_cell(ws.cell(row=legend_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    legend_row += 1
    headers = ['CML', 'Location', 'Nom t (mm)', 'Meas t (mm)', 'Tmin (mm)', 'Comments']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=legend_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)

    # Data rows
    for i in range(15):
        row = legend_row + 1 + i
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            style_cell(cell, font=font_normal, fill=fill_input, border=thin_border, alignment=align_center if col not in [2, 6] else align_left)
        set_row_height(ws, row, 20)

    sig_row = legend_row + 17
    create_signature_block(ws, sig_row, num_cols=6)

    ws.freeze_panes = f'A{legend_row + 1}'
    setup_print_settings(ws, orientation='portrait')

    return ws


def create_shutdown_field_sheet(wb):
    """5. Shutdown Field Sheet (Minimal)"""
    ws = wb.create_sheet("05_Shutdown_Field")

    set_column_widths(ws, {'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 25})

    next_row = create_header_block(ws, "SHUTDOWN FIELD SHEET")

    # Job info
    info_row = next_row
    ws.merge_cells(f'A{info_row}:H{info_row}')
    ws.cell(row=info_row, column=1).value = "JOB INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields = [('Job #:', 'A', 'B'), ('Work Order:', 'C', 'D'), ('Unit/Area:', 'E', 'F'), ('Date:', 'G', 'H')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Data table
    table_row = info_row + 2
    headers = ['CML', 't1 (mm)', 't2 (mm)', 't3 (mm)', 't4 (mm)', 'Min t (mm)', 'Tmin (mm)', 'Comments']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)
    set_row_height(ws, table_row, 28)

    # Data rows with MIN formula
    for i in range(35):
        row = table_row + 1 + i
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if col == 6:  # Min t formula
                cell.value = f'=IF(COUNT(B{row}:E{row})>0,MIN(B{row}:E{row}),"")'
                style_cell(cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
                cell.number_format = '0.00'
            else:
                style_cell(cell, font=font_normal, fill=fill_input, border=thin_border, alignment=align_center if col != 8 else align_left)
        set_row_height(ws, row, 20)

    # Conditional formatting for Min t < Tmin
    add_conditional_formatting_thickness(ws, 6, 7, table_row + 1, table_row + 35)

    ws.freeze_panes = f'A{table_row + 1}'
    setup_print_settings(ws, orientation='portrait')
    ws.print_title_rows = f'1:{table_row}'

    return ws


def create_owner_corrosion_sheet(wb):
    """6. Owner-Mandated Corrosion Sheet"""
    ws = wb.create_sheet("06_Owner_Corrosion")

    set_column_widths(ws, {'A': 12, 'B': 14, 'C': 14, 'D': 14, 'E': 12, 'F': 14, 'G': 14, 'H': 18})

    next_row = create_header_block(ws, "OWNER-MANDATED CORROSION SHEET")

    # Client info
    info_row = next_row
    ws.merge_cells(f'A{info_row}:H{info_row}')
    ws.cell(row=info_row, column=1).value = "CLIENT / UNIT INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields = [('Client:', 'A', 'B'), ('Unit:', 'C', 'D'), ('Risk Class:', 'E', 'F'), ('Analysis Date:', 'G', 'H')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Add Risk Class validation
    risk_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(risk_dv)
    risk_dv.add(ws[f'F{info_row}'])

    # Data table
    table_row = info_row + 2
    headers = ['CML', 'Design t\n(mm)', 'Last t\n(mm)', 'Curr t\n(mm)', '\u0394t\n(mm)', 'CR\n(mm/yr)', 'RL\n(yrs)', 'Next Insp\nDate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)
    set_row_height(ws, table_row, 32)

    # Data rows with formulas
    for i in range(30):
        row = table_row + 1 + i
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if col == 5:  # Delta t = Last t - Curr t
                cell.value = f'=IF(AND(C{row}<>"",D{row}<>""),C{row}-D{row},"")'
                style_cell(cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
                cell.number_format = '0.00'
            elif col == 7:  # RL = (Curr t - Tmin) / CR - simplified without Tmin column
                # We'll leave this as input since Tmin isn't in this template
                style_cell(cell, font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
            else:
                style_cell(cell, font=font_normal, fill=fill_input, border=thin_border, alignment=align_center if col != 8 else align_left)
        set_row_height(ws, row, 20)

    ws.freeze_panes = f'A{table_row + 1}'
    setup_print_settings(ws, orientation='landscape')
    ws.print_title_rows = f'1:{table_row}'

    return ws


def create_pv_shell_grid_sheet(wb):
    """7. Pressure Vessel - Shell Course Grid (API 510)"""
    ws = wb.create_sheet("07_PV_ShellGrid")

    set_column_widths(ws, {'A': 18, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 18})

    next_row = create_header_block(ws, "PRESSURE VESSEL - SHELL COURSE GRID (API 510)")

    # Vessel info
    info_row = next_row
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws.cell(row=info_row, column=1).value = "VESSEL INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields = [('Vessel ID:', 'A', 'B'), ('Tag No:', 'C', 'D'), ('Service:', 'E', 'G')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    info_row += 1
    fields2 = [('Nominal t (mm):', 'A', 'B'), ('Tmin (mm):', 'C', 'D'), ('Material:', 'E', 'G')]
    for label, lc, vc in fields2:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Shell course table
    table_row = info_row + 2
    ws.merge_cells(f'A{table_row}:G{table_row}')
    ws.cell(row=table_row, column=1).value = "SHELL COURSE THICKNESS READINGS (mm)"
    style_cell(ws.cell(row=table_row, column=1), font=font_section, fill=fill_header, border=thin_border, alignment=align_center)

    table_row += 1
    headers = ['Shell Course', 'N (0\u00b0)', 'E (90\u00b0)', 'S (180\u00b0)', 'W (270\u00b0)', 'Minimum', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)

    # Shell course data rows
    for i in range(8):  # 8 shell courses max
        row = table_row + 1 + i
        ws.cell(row=row, column=1).value = f"Course {i + 1}"
        style_cell(ws.cell(row=row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_center)
        for col in range(2, 6):
            style_cell(ws.cell(row=row, column=col), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        # Minimum formula
        min_cell = ws.cell(row=row, column=6)
        min_cell.value = f'=IF(COUNT(B{row}:E{row})>0,MIN(B{row}:E{row}),"")'
        style_cell(min_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        min_cell.number_format = '0.00'
        # Status
        status_cell = ws.cell(row=row, column=7)
        status_cell.value = f'=IF(F{row}="","",IF(F{row}>=$D${info_row},"OK","REVIEW"))'
        style_cell(status_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        set_row_height(ws, row, 22)

    # Overall minimum
    summary_row = table_row + 10
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Overall Minimum Thickness (mm):"
    style_cell(ws.cell(row=summary_row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_right)
    min_all = ws.cell(row=summary_row, column=6)
    min_all.value = f'=IF(COUNT(F{table_row + 1}:F{table_row + 8})>0,MIN(F{table_row + 1}:F{table_row + 8}),"")'
    style_cell(min_all, font=Font(name='Calibri', size=11, bold=True), fill=fill_calculated, border=medium_border, alignment=align_center)
    min_all.number_format = '0.00'

    sig_row = summary_row + 3
    create_signature_block(ws, sig_row, num_cols=7)

    setup_print_settings(ws, orientation='portrait')

    return ws


def create_pv_nozzles_sheet(wb):
    """8. Pressure Vessel - Nozzle Thickness Sheet"""
    ws = wb.create_sheet("08_PV_Nozzles")

    set_column_widths(ws, {'A': 12, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 22})

    next_row = create_header_block(ws, "PRESSURE VESSEL - NOZZLE THICKNESS SHEET")

    # Vessel info
    info_row = next_row
    fields = [('Vessel ID:', 'A', 'B'), ('Tag No:', 'C', 'D'), ('P&ID Ref:', 'E', 'G')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Nozzle table
    table_row = info_row + 2
    headers = ['Nozzle', 'Size/Type', 'Orientation', 'Nom t (mm)', 'Meas t (mm)', 'Tmin (mm)', 'Remarks']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)
    set_row_height(ws, table_row, 28)

    # Common nozzle designations
    nozzle_names = ['N1', 'N2', 'N3', 'N4', 'N5', 'N6', 'N7', 'N8', 'M1', 'M2', 'PSV', 'Drain', 'Vent']
    for i, nozzle in enumerate(nozzle_names):
        row = table_row + 1 + i
        ws.cell(row=row, column=1).value = nozzle
        style_cell(ws.cell(row=row, column=1), font=font_normal, fill=fill_section, border=thin_border, alignment=align_center)
        for col in range(2, 8):
            style_cell(ws.cell(row=row, column=col), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center if col != 7 else align_left)
        set_row_height(ws, row, 20)

    # Extra blank rows
    for i in range(7):
        row = table_row + len(nozzle_names) + 1 + i
        for col in range(1, 8):
            style_cell(ws.cell(row=row, column=col), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center if col != 7 else align_left)
        set_row_height(ws, row, 20)

    # Add conditional formatting
    add_conditional_formatting_thickness(ws, 5, 6, table_row + 1, table_row + len(nozzle_names) + 7)

    ws.freeze_panes = f'A{table_row + 1}'
    setup_print_settings(ws, orientation='portrait')

    return ws


def create_pv_eng_eval_sheet(wb):
    """9. Pressure Vessel - Engineering Evaluation Hybrid"""
    ws = wb.create_sheet("09_PV_EngEval")

    set_column_widths(ws, {'A': 22, 'B': 18, 'C': 18, 'D': 18, 'E': 18})

    next_row = create_header_block(ws, "PRESSURE VESSEL - ENGINEERING EVALUATION")

    # Component info
    info_row = next_row
    ws.merge_cells(f'A{info_row}:E{info_row}')
    ws.cell(row=info_row, column=1).value = "COMPONENT IDENTIFICATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields = [
        ('Vessel ID:', 1), ('Component:', 2), ('Location:', 3),
        ('Measured t (mm):', 4), ('Tmin (mm):', 5)
    ]
    for i, (label, _) in enumerate(fields):
        row = info_row + i
        ws.cell(row=row, column=1).value = label
        style_cell(ws.cell(row=row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        ws.merge_cells(f'B{row}:E{row}')
        style_cell(ws.cell(row=row, column=2), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        set_row_height(ws, row, 22)

    # Engineering notes section
    notes_row = info_row + len(fields) + 1
    ws.merge_cells(f'A{notes_row}:E{notes_row}')
    ws.cell(row=notes_row, column=1).value = "ENGINEERING NOTES"
    style_cell(ws.cell(row=notes_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    note_fields = [
        'Localized thinning observed at:',
        'Assumed corrosion mechanism:',
        'Rate of deterioration assessment:',
        'Remaining service life estimate:',
        'Recommended action:',
    ]
    for i, field in enumerate(note_fields):
        row = notes_row + 1 + i * 2
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row=row, column=1).value = field
        style_cell(ws.cell(row=row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        ws.merge_cells(f'A{row + 1}:E{row + 1}')
        style_cell(ws.cell(row=row + 1, column=1), font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        set_row_height(ws, row, 20)
        set_row_height(ws, row + 1, 30)

    # Acceptability section
    accept_row = notes_row + len(note_fields) * 2 + 2
    ws.merge_cells(f'A{accept_row}:E{accept_row}')
    ws.cell(row=accept_row, column=1).value = "ACCEPTABILITY PER API 510"
    style_cell(ws.cell(row=accept_row, column=1), font=font_section, fill=fill_header, border=thin_border, alignment=align_center)

    accept_row += 1
    ws.cell(row=accept_row, column=1).value = "Determination:"
    style_cell(ws.cell(row=accept_row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
    ws.merge_cells(f'B{accept_row}:C{accept_row}')
    style_cell(ws.cell(row=accept_row, column=2), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)

    # Add validation
    accept_dv = DataValidation(type="list", formula1='"ACCEPTABLE,NOT ACCEPTABLE,REQUIRES FURTHER EVALUATION"', allow_blank=True)
    ws.add_data_validation(accept_dv)
    accept_dv.add(ws.cell(row=accept_row, column=2))

    ws.cell(row=accept_row, column=4).value = "FFS Required:"
    style_cell(ws.cell(row=accept_row, column=4), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
    style_cell(ws.cell(row=accept_row, column=5), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
    ffs_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(ffs_dv)
    ffs_dv.add(ws.cell(row=accept_row, column=5))

    sig_row = accept_row + 3
    create_signature_block(ws, sig_row, num_cols=5)

    setup_print_settings(ws, orientation='portrait')

    return ws


def create_tank_shell_compass_sheet(wb):
    """10. Storage Tank - Shell Compass Table (API 653)"""
    ws = wb.create_sheet("10_Tank_ShellCompass")

    set_column_widths(ws, {'A': 18, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 16})

    next_row = create_header_block(ws, "STORAGE TANK - SHELL COMPASS TABLE (API 653)")

    # Tank info
    info_row = next_row
    ws.merge_cells(f'A{info_row}:G{info_row}')
    ws.cell(row=info_row, column=1).value = "TANK INFORMATION"
    style_cell(ws.cell(row=info_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    info_row += 1
    fields = [('Tank ID:', 'A', 'B'), ('Product:', 'C', 'D'), ('Shell Course:', 'E', 'G')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    info_row += 1
    fields2 = [('Nominal t (mm):', 'A', 'B'), ('Tmin (mm):', 'C', 'D'), ('Height (m):', 'E', 'G')]
    for label, lc, vc in fields2:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Compass readings table
    table_row = info_row + 2
    ws.merge_cells(f'A{table_row}:G{table_row}')
    ws.cell(row=table_row, column=1).value = "COMPASS THICKNESS READINGS (mm)"
    style_cell(ws.cell(row=table_row, column=1), font=font_section, fill=fill_header, border=thin_border, alignment=align_center)

    table_row += 1
    headers = ['Position', '0\u00b0 (N)', '90\u00b0 (E)', '180\u00b0 (S)', '270\u00b0 (W)', 'Minimum', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)

    # Position rows (Top, Middle, Bottom of course)
    positions = ['Top (6" from seam)', 'Middle', 'Bottom (6" from seam)']
    for i, pos in enumerate(positions):
        row = table_row + 1 + i
        ws.cell(row=row, column=1).value = pos
        style_cell(ws.cell(row=row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        for col in range(2, 6):
            style_cell(ws.cell(row=row, column=col), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        # Min formula
        min_cell = ws.cell(row=row, column=6)
        min_cell.value = f'=IF(COUNT(B{row}:E{row})>0,MIN(B{row}:E{row}),"")'
        style_cell(min_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        min_cell.number_format = '0.00'
        # Status
        status_cell = ws.cell(row=row, column=7)
        status_cell.value = f'=IF(F{row}="","",IF(F{row}>=$D${info_row},"OK","REVIEW"))'
        style_cell(status_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        set_row_height(ws, row, 24)

    # Overall summary
    summary_row = table_row + 5
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Minimum Thickness This Course (mm):"
    style_cell(ws.cell(row=summary_row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_right)
    min_course = ws.cell(row=summary_row, column=6)
    min_course.value = f'=IF(COUNT(F{table_row + 1}:F{table_row + 3})>0,MIN(F{table_row + 1}:F{table_row + 3}),"")'
    style_cell(min_course, font=Font(name='Calibri', size=11, bold=True), fill=fill_calculated, border=medium_border, alignment=align_center)
    min_course.number_format = '0.00'

    sig_row = summary_row + 3
    create_signature_block(ws, sig_row, num_cols=7)

    setup_print_settings(ws, orientation='portrait')

    return ws


def create_tank_multi_course_sheet(wb):
    """11. Storage Tank - Multi-Course Summary"""
    ws = wb.create_sheet("11_Tank_MultiCourse")

    set_column_widths(ws, {'A': 14, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 18})

    next_row = create_header_block(ws, "STORAGE TANK - MULTI-COURSE SUMMARY")

    # Tank info
    info_row = next_row
    fields = [('Tank ID:', 'A', 'B'), ('Product:', 'C', 'D'), ('Inspection Date:', 'E', 'G')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'G':
            ws.merge_cells(f'F{info_row}:G{info_row}')
            style_cell(ws[f'F{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Summary table
    table_row = info_row + 2
    headers = ['Course', 'Nom t (mm)', 'Min t (mm)', 'Avg t (mm)', 'Tmin (mm)', '% Loss', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col)
        cell.value = header
        style_cell(cell, font=font_header, fill=fill_header, border=medium_border, alignment=align_center)
    set_row_height(ws, table_row, 28)

    # Course rows
    for i in range(12):  # Up to 12 courses
        row = table_row + 1 + i
        ws.cell(row=row, column=1).value = i + 1
        style_cell(ws.cell(row=row, column=1), font=font_normal, fill=fill_section, border=thin_border, alignment=align_center)
        for col in range(2, 6):
            style_cell(ws.cell(row=row, column=col), font=font_normal, fill=fill_input, border=thin_border, alignment=align_center)
        # % Loss formula
        loss_cell = ws.cell(row=row, column=6)
        loss_cell.value = f'=IF(AND(B{row}<>"",C{row}<>""),ROUND((B{row}-C{row})/B{row}*100,1),"")'
        style_cell(loss_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        loss_cell.number_format = '0.0"%"'
        # Status
        status_cell = ws.cell(row=row, column=7)
        status_cell.value = f'=IF(OR(C{row}="",E{row}=""),"",IF(C{row}>=E{row},"OK","BELOW Tmin"))'
        style_cell(status_cell, font=font_normal, fill=fill_calculated, border=thin_border, alignment=align_center)
        set_row_height(ws, row, 20)

    # Overall summary
    summary_row = table_row + 14
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Overall Min (mm):"
    style_cell(ws.cell(row=summary_row, column=1), font=font_label, fill=fill_section, border=thin_border, alignment=align_right)
    overall_min = ws.cell(row=summary_row, column=3)
    overall_min.value = f'=IF(COUNT(C{table_row + 1}:C{table_row + 12})>0,MIN(C{table_row + 1}:C{table_row + 12}),"")'
    style_cell(overall_min, font=Font(name='Calibri', size=11, bold=True), fill=fill_calculated, border=medium_border, alignment=align_center)
    overall_min.number_format = '0.00'

    ws.cell(row=summary_row, column=4).value = "Max % Loss:"
    style_cell(ws.cell(row=summary_row, column=4), font=font_label, fill=fill_section, border=thin_border, alignment=align_right)
    max_loss = ws.cell(row=summary_row, column=6)
    max_loss.value = f'=IF(COUNT(F{table_row + 1}:F{table_row + 12})>0,MAX(F{table_row + 1}:F{table_row + 12}),"")'
    style_cell(max_loss, font=Font(name='Calibri', size=11, bold=True), fill=fill_calculated, border=medium_border, alignment=align_center)
    max_loss.number_format = '0.0"%"'

    sig_row = summary_row + 3
    create_signature_block(ws, sig_row, num_cols=7)

    ws.freeze_panes = f'A{table_row + 1}'
    setup_print_settings(ws, orientation='portrait')

    return ws


def create_tank_inspector_notes_sheet(wb):
    """12. Storage Tank - Inspector Notebook Page"""
    ws = wb.create_sheet("12_Tank_InspectorNotes")

    set_column_widths(ws, {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20})

    next_row = create_header_block(ws, "INSPECTOR NOTEBOOK PAGE")

    # Basic info
    info_row = next_row
    fields = [('Tank/Equipment ID:', 'A', 'B'), ('Date:', 'C', 'C'), ('Page:', 'D', 'E')]
    for label, lc, vc in fields:
        ws[f'{lc}{info_row}'] = label
        style_cell(ws[f'{lc}{info_row}'], font=font_label, fill=fill_section, border=thin_border, alignment=align_left)
        if vc == 'E':
            style_cell(ws[f'E{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        elif vc == 'C':
            style_cell(ws[f'C{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)
        else:
            style_cell(ws[f'{vc}{info_row}'], font=font_normal, fill=fill_input, border=thin_border, alignment=align_left)

    # Free-form notes area
    notes_row = info_row + 2
    ws.merge_cells(f'A{notes_row}:E{notes_row}')
    ws.cell(row=notes_row, column=1).value = "FREE-FORM NOTES / SKETCHES"
    style_cell(ws.cell(row=notes_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    # Large input area
    notes_row += 1
    ws.merge_cells(f'A{notes_row}:E{notes_row + 30}')
    notes_cell = ws.cell(row=notes_row, column=1)
    notes_cell.value = ""
    style_cell(notes_cell, font=font_normal, fill=fill_input, border=medium_border, alignment=Alignment(horizontal='left', vertical='top', wrap_text=True))

    for r in range(notes_row, notes_row + 31):
        set_row_height(ws, r, 18)

    # Quick reference box
    ref_row = notes_row + 32
    ws.merge_cells(f'A{ref_row}:E{ref_row}')
    ws.cell(row=ref_row, column=1).value = "QUICK REFERENCE - COMMON ANNOTATIONS"
    style_cell(ws.cell(row=ref_row, column=1), font=font_section, fill=fill_section, border=thin_border, alignment=align_center)

    ref_row += 1
    ws.merge_cells(f'A{ref_row}:E{ref_row + 2}')
    ref_cell = ws.cell(row=ref_row, column=1)
    ref_cell.value = "\u2022 Circle critical readings  \u2022 Arrow to indicate problem areas  \u2022 'X' marks defects\n\u2022 Use dashed lines for estimated boundaries  \u2022 Note CML numbers clearly\n\u2022 Include dimensions and distances where relevant"
    style_cell(ref_cell, font=font_small, fill=fill_section, border=thin_border, alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))

    setup_print_settings(ws, orientation='portrait')

    return ws


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Generate the complete professional UT CML Templates workbook"""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create all template sheets
    print("Creating professional UT CML Templates...")

    create_piping_circuit_sheet(wb)
    print("  [1/12] Piping Circuit Thickness Sheet")

    create_orientation_grid_sheet(wb)
    print("  [2/12] Orientation-Based CML Grid")

    create_pipe_shoe_cui_sheet(wb)
    print("  [3/12] Pipe Shoe / CUI Thickness Sheet")

    create_isometric_legend_sheet(wb)
    print("  [4/12] Isometric + CML Legend")

    create_shutdown_field_sheet(wb)
    print("  [5/12] Shutdown Field Sheet")

    create_owner_corrosion_sheet(wb)
    print("  [6/12] Owner-Mandated Corrosion Sheet")

    create_pv_shell_grid_sheet(wb)
    print("  [7/12] PV Shell Course Grid")

    create_pv_nozzles_sheet(wb)
    print("  [8/12] PV Nozzle Thickness Sheet")

    create_pv_eng_eval_sheet(wb)
    print("  [9/12] PV Engineering Evaluation")

    create_tank_shell_compass_sheet(wb)
    print("  [10/12] Tank Shell Compass Table")

    create_tank_multi_course_sheet(wb)
    print("  [11/12] Tank Multi-Course Summary")

    create_tank_inspector_notes_sheet(wb)
    print("  [12/12] Inspector Notebook Page")

    # Save workbook
    output_path = 'UT_CML_Templates_Professional.xlsx'
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()

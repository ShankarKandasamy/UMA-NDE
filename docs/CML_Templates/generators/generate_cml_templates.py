"""
Generate realistic CML (Condition Monitoring Location) templates and sample data
for UT thickness surveys across piping, pressure vessels, and storage tanks.

These templates represent typical field inspection forms used across North American
oil & gas, petrochemical, and industrial facilities.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta

# Common styles
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
SUBHEADER_FILL = PatternFill('solid', fgColor='2E75B6')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=9)
DATA_FONT = Font(name='Arial', size=9)
TITLE_FONT = Font(bold=True, name='Arial', size=14)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

# Alert colors
ALERT_FILL = PatternFill('solid', fgColor='FFCCCC')  # Red for below min
WARNING_FILL = PatternFill('solid', fgColor='FFFFCC')  # Yellow for approaching

def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def apply_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

def apply_subheader_style(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

def apply_data_style(cell, center=True):
    cell.font = DATA_FONT
    cell.alignment = CENTER if center else LEFT
    cell.border = THIN_BORDER

# ============================================================================
# PIPING CML TEMPLATES (API 570)
# ============================================================================

def create_piping_template_style1(wb):
    """Style 1: Traditional tabular format - most common"""
    ws = wb.create_sheet("Piping_Style1_Template")
    
    # Header section
    ws.merge_cells('A1:L1')
    ws['A1'] = 'UT THICKNESS SURVEY - PIPING'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    # Info fields
    info_fields = [
        ('A3', 'Circuit ID:', 'B3', ''),
        ('A4', 'Service:', 'B4', ''),
        ('A5', 'P&ID No.:', 'B5', ''),
        ('A6', 'Material:', 'B6', ''),
        ('D3', 'Inspector:', 'E3', ''),
        ('D4', 'Cert Level:', 'E4', ''),
        ('D5', 'Date:', 'E5', ''),
        ('D6', 'Instrument:', 'E6', ''),
        ('G3', 'Design Press:', 'H3', ''),
        ('G4', 'Design Temp:', 'H4', ''),
        ('G5', 'Cal Block:', 'H5', ''),
        ('G6', 'Cal Due:', 'H6', ''),
    ]
    
    for label_cell, label, value_cell, value in info_fields:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell] = value
        ws[value_cell].border = Border(bottom=Side(style='thin'))
    
    # Data table headers - Row 8
    headers = ['CML', 'Component', 'NPS', 'Sched', 'Nominal\n(mm)', 't-min\n(mm)', 
               'Reading 1', 'Reading 2', 'Reading 3', 'Reading 4', 'Min\n(mm)', 
               'Corr Rate\n(mm/yr)', 'Remaining\nLife (yr)', 'Prev Reading\n(mm)', 
               'Prev Date', 'Condition', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[8].height = 40
    
    # Set column widths
    widths = {'A': 8, 'B': 14, 'C': 6, 'D': 6, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 
              'I': 10, 'J': 10, 'K': 10, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 25}
    set_column_widths(ws, widths)
    
    # Empty data rows
    for row in range(9, 34):
        for col in range(1, 18):
            cell = ws.cell(row=row, column=col, value='')
            apply_data_style(cell)
    
    return ws


def create_piping_template_style2(wb):
    """Style 2: Grid-focused format for detailed CML mapping"""
    ws = wb.create_sheet("Piping_Style2_Grid")
    
    ws.merge_cells('A1:N1')
    ws['A1'] = 'PIPING THICKNESS SURVEY - GRID FORMAT'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    # Simplified header
    headers_row3 = ['Circuit:', '', 'Date:', '', 'Inspector:', '', 'Instrument:', '']
    for col, val in enumerate(headers_row3, 1):
        ws.cell(row=3, column=col, value=val)
    
    # This format has one CML per section with grid readings
    ws['A5'] = 'CML:'
    ws['B5'] = ''
    ws['C5'] = 'Component:'
    ws['D5'] = ''
    ws['E5'] = 'Size:'
    ws['F5'] = ''
    ws['G5'] = 'Nominal:'
    ws['H5'] = ''
    
    # Grid headers
    ws['A7'] = 'Grid'
    for col, pos in enumerate(['1', '2', '3', '4'], 2):
        ws.cell(row=7, column=col, value=pos)
        apply_subheader_style(ws.cell(row=7, column=col))
    
    apply_subheader_style(ws['A7'])
    
    # Grid rows
    for row, letter in enumerate(['A', 'B', 'C', 'D'], 8):
        ws.cell(row=row, column=1, value=letter)
        apply_subheader_style(ws.cell(row=row, column=1))
        for col in range(2, 6):
            apply_data_style(ws.cell(row=row, column=col))
    
    # Summary
    ws['A13'] = 'Minimum:'
    ws['A14'] = 'Location:'
    ws['A15'] = 'Condition:'
    
    ws.column_dimensions['A'].width = 12
    for col in 'BCDE':
        ws.column_dimensions[col].width = 10
    
    return ws


def create_piping_sample_data(wb):
    """Create filled example with realistic piping data"""
    ws = wb.create_sheet("Piping_Sample_FW101")
    
    # Header
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'UT THICKNESS SURVEY - PIPING'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    # Filled info
    info = [
        ('A3', 'Circuit ID:', 'B3', 'FW-101'),
        ('A4', 'Service:', 'B4', 'Boiler Feedwater'),
        ('A5', 'P&ID No.:', 'B5', 'P-101-01-A'),
        ('A6', 'Material:', 'B6', 'A106 Gr.B'),
        ('D3', 'Inspector:', 'E3', 'J. Smith'),
        ('D4', 'Cert Level:', 'E4', 'ASNT UT Level II'),
        ('D5', 'Date:', 'E5', '2026-01-21'),
        ('D6', 'Instrument:', 'E6', 'Olympus 38DL+ S/N 12345'),
        ('G3', 'Design Press:', 'H3', '1500 psig'),
        ('G4', 'Design Temp:', 'H4', '450°F'),
        ('G5', 'Cal Block:', 'H5', 'CB-001'),
        ('G6', 'Cal Due:', 'H6', '2026-06-15'),
    ]
    
    for label_cell, label, value_cell, value in info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell] = value
        ws[value_cell].font = DATA_FONT
        ws[value_cell].border = Border(bottom=Side(style='thin'))
    
    # Headers
    headers = ['CML', 'Component', 'NPS', 'Sched', 'Nominal\n(mm)', 't-min\n(mm)', 
               'Reading 1', 'Reading 2', 'Reading 3', 'Reading 4', 'Min\n(mm)', 
               'Corr Rate\n(mm/yr)', 'Remaining\nLife (yr)', 'Prev Reading\n(mm)', 
               'Prev Date', 'Condition', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[8].height = 40
    
    # Sample data - realistic feedwater piping circuit
    piping_data = [
        ['001', '90° Elbow', '6', '80', 7.11, 3.56, 6.82, 6.91, 6.78, 6.85, 6.78, 0.08, 40.3, 6.95, '2024-01-15', 'Good', ''],
        ['002', 'Straight', '6', '80', 7.11, 3.56, 6.95, 6.98, '', '', 6.95, 0.05, 67.8, 7.02, '2024-01-15', 'Good', ''],
        ['003', 'Tee-Run', '6', '80', 7.11, 3.56, 6.88, 6.75, 6.82, 6.79, 6.75, 0.12, 26.6, 6.92, '2024-01-15', 'Good', 'Slight erosion at branch'],
        ['004', 'Tee-Branch', '4', '80', 6.02, 3.01, 5.85, 5.92, 5.78, 5.88, 5.78, 0.09, 30.8, 5.95, '2024-01-15', 'Good', ''],
        ['005', 'Reducer', '6x4', '80', 7.11, 3.56, 6.91, 6.88, 6.95, '', 6.88, 0.06, 55.3, 6.98, '2024-01-15', 'Good', ''],
        ['006', 'Straight', '4', '80', 6.02, 3.01, 5.92, 5.88, '', '', 5.88, 0.07, 41.0, 5.98, '2024-01-15', 'Good', ''],
        ['007', '45° Elbow', '4', '80', 6.02, 3.01, 5.78, 5.82, 5.75, 5.80, 5.75, 0.11, 24.9, 5.88, '2024-01-15', 'Fair', 'Monitor - higher rate'],
        ['008', 'Straight', '4', '80', 6.02, 3.01, 5.95, 5.91, '', '', 5.91, 0.05, 58.0, 5.98, '2024-01-15', 'Good', ''],
        ['009', '90° Elbow', '4', '80', 6.02, 3.01, 4.52, 4.48, 4.61, 4.55, 4.48, 0.25, 5.9, 4.85, '2024-01-15', 'ALERT', '⚠️ BELOW ALERT - Schedule repair'],
        ['010', 'Straight', '4', '80', 6.02, 3.01, 5.88, 5.85, '', '', 5.85, 0.06, 47.3, 5.92, '2024-01-15', 'Good', ''],
        ['011', 'Control Valve', '4', '80', 6.02, 3.01, 5.75, 5.72, 5.78, 5.71, 5.71, 0.10, 27.0, 5.82, '2024-01-15', 'Fair', 'Downstream of CV'],
        ['012', 'Straight', '4', '80', 6.02, 3.01, 5.91, 5.88, '', '', 5.88, 0.05, 57.4, 5.95, '2024-01-15', 'Good', ''],
    ]
    
    for row_idx, row_data in enumerate(piping_data, 9):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, center=(col_idx != 17))
            
            # Highlight alert rows
            if row_data[15] == 'ALERT':
                cell.fill = ALERT_FILL
            elif row_data[15] == 'Fair':
                cell.fill = WARNING_FILL
    
    # Column widths
    widths = {'A': 8, 'B': 14, 'C': 6, 'D': 6, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 
              'I': 10, 'J': 10, 'K': 10, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 10, 'Q': 30}
    set_column_widths(ws, widths)
    
    # Footer notes
    ws['A23'] = 'NOTES:'
    ws['A23'].font = Font(bold=True, name='Arial', size=9)
    ws['A24'] = '- CML 009 requires immediate attention. Wall thickness approaching t-min.'
    ws['A25'] = '- CML 007 and 011 showing elevated corrosion rates. Recommend 1-year re-inspection interval.'
    ws['A26'] = '- All readings taken using 5MHz dual element probe, V-path calibration verified on 4-step block.'
    
    return ws


def create_piping_sample2(wb):
    """Create second piping example - different circuit"""
    ws = wb.create_sheet("Piping_Sample_CW205")
    
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'UT THICKNESS SURVEY - PIPING'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    info = [
        ('A3', 'Circuit ID:', 'B3', 'CW-205'),
        ('A4', 'Service:', 'B4', 'Cooling Water Supply'),
        ('A5', 'P&ID No.:', 'B5', 'P-205-03-B'),
        ('A6', 'Material:', 'B6', 'A53 Gr.B (Lined)'),
        ('D3', 'Inspector:', 'E3', 'M. Johnson'),
        ('D4', 'Cert Level:', 'E4', 'CGSB UT Level II'),
        ('D5', 'Date:', 'E5', '2026-01-20'),
        ('D6', 'Instrument:', 'E6', 'GE DMS Go S/N 98765'),
        ('G3', 'Design Press:', 'H3', '150 psig'),
        ('G4', 'Design Temp:', 'H4', '120°F'),
        ('G5', 'Cal Block:', 'H5', 'CB-002'),
        ('G6', 'Cal Due:', 'H6', '2026-04-30'),
    ]
    
    for label_cell, label, value_cell, value in info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell] = value
        ws[value_cell].font = DATA_FONT
    
    headers = ['CML', 'Component', 'NPS', 'Sched', 'Nominal\n(mm)', 't-min\n(mm)', 
               'Reading 1', 'Reading 2', 'Reading 3', 'Reading 4', 'Min\n(mm)', 
               'Corr Rate\n(mm/yr)', 'Remaining\nLife (yr)', 'Prev Reading\n(mm)', 
               'Prev Date', 'Condition', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[8].height = 40
    
    # Cooling water - larger pipe, typically more corrosion
    cw_data = [
        ['001', '90° Elbow', '12', '40', 10.31, 5.16, 8.85, 8.92, 8.78, 8.88, 8.78, 0.18, 20.1, 9.12, '2023-06-10', 'Fair', 'ID pitting observed'],
        ['002', 'Straight', '12', '40', 10.31, 5.16, 9.45, 9.52, '', '', 9.45, 0.12, 35.8, 9.62, '2023-06-10', 'Good', ''],
        ['003', 'Straight', '12', '40', 10.31, 5.16, 9.38, 9.41, '', '', 9.38, 0.14, 30.1, 9.58, '2023-06-10', 'Good', ''],
        ['004', 'Tee-Run', '12', '40', 10.31, 5.16, 9.15, 9.08, 9.22, 9.11, 9.08, 0.16, 24.5, 9.32, '2023-06-10', 'Good', ''],
        ['005', 'Tee-Branch', '8', '40', 8.18, 4.09, 7.52, 7.48, 7.55, 7.45, 7.45, 0.15, 22.4, 7.68, '2023-06-10', 'Good', ''],
        ['006', '90° Elbow', '12', '40', 10.31, 5.16, 8.42, 8.38, 8.51, 8.35, 8.35, 0.22, 14.5, 8.68, '2023-06-10', 'Fair', 'Extrados thinning'],
        ['007', 'Straight', '12', '40', 10.31, 5.16, 9.28, 9.35, '', '', 9.28, 0.13, 31.7, 9.48, '2023-06-10', 'Good', ''],
        ['008', 'Reducer', '12x8', '40', 10.31, 5.16, 9.18, 9.25, 9.12, '', 9.12, 0.14, 28.3, 9.32, '2023-06-10', 'Good', ''],
    ]
    
    for row_idx, row_data in enumerate(cw_data, 9):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, center=(col_idx != 17))
            if row_data[15] == 'Fair':
                cell.fill = WARNING_FILL
    
    widths = {'A': 8, 'B': 14, 'C': 6, 'D': 6, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 
              'I': 10, 'J': 10, 'K': 10, 'L': 12, 'M': 12, 'N': 12, 'O': 12, 'P': 10, 'Q': 25}
    set_column_widths(ws, widths)
    
    return ws


# ============================================================================
# PRESSURE VESSEL CML TEMPLATES (API 510)
# ============================================================================

def create_vessel_template(wb):
    """Pressure vessel thickness survey template"""
    ws = wb.create_sheet("Vessel_Template")
    
    ws.merge_cells('A1:N1')
    ws['A1'] = 'PRESSURE VESSEL THICKNESS SURVEY'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    # Vessel identification section
    ws['A3'] = 'VESSEL IDENTIFICATION'
    ws['A3'].font = Font(bold=True, name='Arial', size=11)
    
    vessel_info = [
        ('A4', 'Vessel Tag:', 'B4', ''),
        ('A5', 'Service:', 'B5', ''),
        ('A6', 'Manufacturer:', 'B6', ''),
        ('A7', 'Year Built:', 'B7', ''),
        ('D4', 'National Board No:', 'E4', ''),
        ('D5', 'MAWP:', 'E5', ''),
        ('D6', 'Design Temp:', 'E6', ''),
        ('D7', 'Material:', 'E7', ''),
        ('G4', 'Corr. Allowance:', 'H4', ''),
        ('G5', 'Joint Efficiency:', 'H5', ''),
        ('G6', 'Last Internal:', 'H6', ''),
        ('G7', 'Next Internal Due:', 'H7', ''),
    ]
    
    for label_cell, label, value_cell, value in vessel_info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell].border = Border(bottom=Side(style='thin'))
    
    # Inspection info
    ws['A9'] = 'INSPECTION DETAILS'
    ws['A9'].font = Font(bold=True, name='Arial', size=11)
    
    insp_info = [
        ('A10', 'Inspector:', 'B10', ''),
        ('A11', 'Date:', 'B11', ''),
        ('D10', 'Instrument:', 'E10', ''),
        ('D11', 'Procedure:', 'E11', ''),
        ('G10', 'Inspection Type:', 'H10', ''),
        ('G11', 'Cal Block:', 'H11', ''),
    ]
    
    for label_cell, label, value_cell, value in insp_info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell].border = Border(bottom=Side(style='thin'))
    
    # Data table
    ws['A13'] = 'THICKNESS READINGS'
    ws['A13'].font = Font(bold=True, name='Arial', size=11)
    
    headers = ['Zone', 'CML', 'Orientation', 'Elevation', 'Nominal\n(mm)', 't-min\n(mm)',
               'Reading', 'Prev\n(mm)', 'Prev Date', 'Corr Rate\n(mm/yr)', 'Remaining\nLife (yr)', 
               'Condition', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[14].height = 35
    
    # Empty rows with zone labels
    zones = ['TOP HEAD', '', '', 'SHELL COURSE 1', '', '', '', 'SHELL COURSE 2', '', '', '',
             'SHELL COURSE 3', '', '', '', 'BOTTOM HEAD', '', '', 'NOZZLE N1', 'NOZZLE N2', 'NOZZLE MH-1']
    
    for row_idx, zone in enumerate(zones, 15):
        ws.cell(row=row_idx, column=1, value=zone)
        for col in range(1, 14):
            apply_data_style(ws.cell(row=row_idx, column=col))
    
    widths = {'A': 16, 'B': 8, 'C': 12, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 
              'H': 10, 'I': 12, 'J': 12, 'K': 12, 'L': 10, 'M': 25}
    set_column_widths(ws, widths)
    
    return ws


def create_vessel_sample(wb):
    """Filled pressure vessel example"""
    ws = wb.create_sheet("Vessel_Sample_V101")
    
    ws.merge_cells('A1:N1')
    ws['A1'] = 'PRESSURE VESSEL THICKNESS SURVEY'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    # Filled vessel info
    vessel_info = [
        ('A4', 'Vessel Tag:', 'B4', 'V-101'),
        ('A5', 'Service:', 'B5', 'Amine Contactor'),
        ('A6', 'Manufacturer:', 'B6', 'Nooter Corp'),
        ('A7', 'Year Built:', 'B7', '1998'),
        ('D4', 'National Board No:', 'E4', 'NB-45678'),
        ('D5', 'MAWP:', 'E5', '250 psig'),
        ('D6', 'Design Temp:', 'E6', '300°F'),
        ('D7', 'Material:', 'E7', 'SA-516-70'),
        ('G4', 'Corr. Allowance:', 'H4', '3.175 mm (1/8")'),
        ('G5', 'Joint Efficiency:', 'H5', '1.0'),
        ('G6', 'Last Internal:', 'H6', '2022-04-15'),
        ('G7', 'Next Internal Due:', 'H7', '2027-04-15'),
    ]
    
    ws['A3'] = 'VESSEL IDENTIFICATION'
    ws['A3'].font = Font(bold=True, name='Arial', size=11)
    
    for label_cell, label, value_cell, value in vessel_info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell] = value
        ws[value_cell].font = DATA_FONT
    
    # Inspection info
    ws['A9'] = 'INSPECTION DETAILS'
    ws['A9'].font = Font(bold=True, name='Arial', size=11)
    
    insp_info = [
        ('A10', 'Inspector:', 'B10', 'R. Williams'),
        ('A11', 'Date:', 'B11', '2026-01-18'),
        ('D10', 'Instrument:', 'E10', 'Olympus 38DL+ S/N 23456'),
        ('D11', 'Procedure:', 'E11', 'UT-PROC-001 Rev.5'),
        ('G10', 'Inspection Type:', 'H10', 'External On-Stream'),
        ('G11', 'Cal Block:', 'H11', 'CB-003'),
    ]
    
    for label_cell, label, value_cell, value in insp_info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell] = value
        ws[value_cell].font = DATA_FONT
    
    ws['A13'] = 'THICKNESS READINGS'
    ws['A13'].font = Font(bold=True, name='Arial', size=11)
    
    headers = ['Zone', 'CML', 'Orientation', 'Elevation', 'Nominal\n(mm)', 't-min\n(mm)',
               'Reading', 'Prev\n(mm)', 'Prev Date', 'Corr Rate\n(mm/yr)', 'Remaining\nLife (yr)', 
               'Condition', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[14].height = 35
    
    # Sample vessel data
    vessel_data = [
        ['TOP HEAD', 'TH-1', 'Center', '-', 19.05, 9.53, 18.21, 18.45, '2022-04-15', 0.06, 145, 'Good', ''],
        ['TOP HEAD', 'TH-2', 'N', '-', 19.05, 9.53, 18.15, 18.38, '2022-04-15', 0.06, 143, 'Good', ''],
        ['TOP HEAD', 'TH-3', 'S', '-', 19.05, 9.53, 18.28, 18.52, '2022-04-15', 0.06, 146, 'Good', ''],
        ['SHELL 1', 'SC1-N', 'N', 'Mid', 25.40, 12.70, 23.85, 24.12, '2022-04-15', 0.07, 159, 'Good', ''],
        ['SHELL 1', 'SC1-E', 'E', 'Mid', 25.40, 12.70, 23.62, 23.88, '2022-04-15', 0.07, 156, 'Good', ''],
        ['SHELL 1', 'SC1-S', 'S', 'Mid', 25.40, 12.70, 23.78, 24.05, '2022-04-15', 0.07, 158, 'Good', ''],
        ['SHELL 1', 'SC1-W', 'W', 'Mid', 25.40, 12.70, 23.91, 24.18, '2022-04-15', 0.07, 160, 'Good', ''],
        ['SHELL 2', 'SC2-N', 'N', 'Mid', 25.40, 12.70, 22.45, 22.78, '2022-04-15', 0.09, 108, 'Good', 'Liquid level zone'],
        ['SHELL 2', 'SC2-E', 'E', 'Mid', 25.40, 12.70, 21.82, 22.18, '2022-04-15', 0.10, 91, 'Fair', 'Higher corrosion'],
        ['SHELL 2', 'SC2-S', 'S', 'Mid', 25.40, 12.70, 22.38, 22.72, '2022-04-15', 0.09, 108, 'Good', ''],
        ['SHELL 2', 'SC2-W', 'W', 'Mid', 25.40, 12.70, 22.52, 22.85, '2022-04-15', 0.09, 109, 'Good', ''],
        ['SHELL 3', 'SC3-N', 'N', 'Mid', 25.40, 12.70, 23.95, 24.22, '2022-04-15', 0.07, 161, 'Good', ''],
        ['SHELL 3', 'SC3-E', 'E', 'Mid', 25.40, 12.70, 24.02, 24.28, '2022-04-15', 0.07, 162, 'Good', ''],
        ['SHELL 3', 'SC3-S', 'S', 'Mid', 25.40, 12.70, 23.88, 24.15, '2022-04-15', 0.07, 160, 'Good', ''],
        ['SHELL 3', 'SC3-W', 'W', 'Mid', 25.40, 12.70, 24.08, 24.35, '2022-04-15', 0.07, 162, 'Good', ''],
        ['BTM HEAD', 'BH-1', 'Center', '-', 19.05, 9.53, 17.52, 17.85, '2022-04-15', 0.09, 89, 'Fair', 'Bottom sludge zone'],
        ['BTM HEAD', 'BH-2', 'N', '-', 19.05, 9.53, 17.88, 18.18, '2022-04-15', 0.08, 105, 'Good', ''],
        ['BTM HEAD', 'BH-3', 'S', '-', 19.05, 9.53, 17.75, 18.05, '2022-04-15', 0.08, 103, 'Good', ''],
        ['NOZZLE', 'N1-Neck', 'Inlet 6"', '-', 11.13, 5.57, 10.52, 10.72, '2022-04-15', 0.05, 99, 'Good', ''],
        ['NOZZLE', 'N2-Neck', 'Outlet 6"', '-', 11.13, 5.57, 10.48, 10.68, '2022-04-15', 0.05, 98, 'Good', ''],
        ['NOZZLE', 'MH-Neck', 'Manway 18"', '-', 12.70, 6.35, 12.15, 12.32, '2022-04-15', 0.05, 116, 'Good', ''],
    ]
    
    for row_idx, row_data in enumerate(vessel_data, 15):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell, center=(col_idx != 13))
            if 'Fair' in str(row_data[11]):
                cell.fill = WARNING_FILL
    
    widths = {'A': 12, 'B': 10, 'C': 12, 'D': 8, 'E': 10, 'F': 10, 'G': 10, 
              'H': 10, 'I': 12, 'J': 12, 'K': 12, 'L': 10, 'M': 25}
    set_column_widths(ws, widths)
    
    return ws


# ============================================================================
# STORAGE TANK CML TEMPLATES (API 653)
# ============================================================================

def create_tank_template(wb):
    """Storage tank thickness survey template"""
    ws = wb.create_sheet("Tank_Template")
    
    ws.merge_cells('A1:L1')
    ws['A1'] = 'ABOVEGROUND STORAGE TANK THICKNESS SURVEY'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    ws['A3'] = 'TANK IDENTIFICATION'
    ws['A3'].font = Font(bold=True, name='Arial', size=11)
    
    tank_info = [
        ('A4', 'Tank No:', 'B4', ''),
        ('A5', 'Service:', 'B5', ''),
        ('A6', 'Capacity:', 'B6', ''),
        ('A7', 'Diameter:', 'B7', ''),
        ('D4', 'Height:', 'E4', ''),
        ('D5', 'Year Built:', 'E5', ''),
        ('D6', 'Roof Type:', 'E6', ''),
        ('D7', 'Floor Type:', 'E7', ''),
        ('G4', 'Shell Material:', 'H4', ''),
        ('G5', 'Floor Material:', 'H5', ''),
        ('G6', 'Last Inspection:', 'H6', ''),
        ('G7', 'Inspection Type:', 'H7', ''),
    ]
    
    for label_cell, label, value_cell, value in tank_info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell].border = Border(bottom=Side(style='thin'))
    
    # Shell section
    ws['A10'] = 'SHELL THICKNESS READINGS (mm)'
    ws['A10'].font = Font(bold=True, name='Arial', size=11)
    
    shell_headers = ['Course', 'Original\nThk (mm)', 'N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW', 'Minimum', 't-min\n(mm)']
    
    for col, header in enumerate(shell_headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[11].height = 30
    
    for row in range(12, 20):
        ws.cell(row=row, column=1, value=row - 11)
        for col in range(1, 13):
            apply_data_style(ws.cell(row=row, column=col))
    
    # Floor section
    ws['A22'] = 'FLOOR THICKNESS READINGS (mm) - See attached sketch for grid reference'
    ws['A22'].font = Font(bold=True, name='Arial', size=11)
    
    floor_headers = ['Grid', '1', '2', '3', '4', '5', '6', '7', '8', 'Min']
    for col, header in enumerate(floor_headers, 1):
        cell = ws.cell(row=23, column=col, value=header)
        apply_subheader_style(cell)
    
    for row, letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], 24):
        ws.cell(row=row, column=1, value=letter)
        for col in range(1, 11):
            apply_data_style(ws.cell(row=row, column=col))
    
    # Annular plate section
    ws['A34'] = 'ANNULAR PLATE READINGS (mm)'
    ws['A34'].font = Font(bold=True, name='Arial', size=11)
    
    widths = {'A': 10, 'B': 12, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8, 'I': 8, 'J': 8, 'K': 10, 'L': 10}
    set_column_widths(ws, widths)
    
    return ws


def create_tank_sample(wb):
    """Filled storage tank example"""
    ws = wb.create_sheet("Tank_Sample_T501")
    
    ws.merge_cells('A1:L1')
    ws['A1'] = 'ABOVEGROUND STORAGE TANK THICKNESS SURVEY'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER
    
    ws['A3'] = 'TANK IDENTIFICATION'
    ws['A3'].font = Font(bold=True, name='Arial', size=11)
    
    tank_info = [
        ('A4', 'Tank No:', 'B4', 'T-501'),
        ('A5', 'Service:', 'B5', 'Crude Oil Storage'),
        ('A6', 'Capacity:', 'B6', '50,000 BBL'),
        ('A7', 'Diameter:', 'B7', '120 ft (36.6m)'),
        ('D4', 'Height:', 'E4', '48 ft (14.6m)'),
        ('D5', 'Year Built:', 'E5', '1985'),
        ('D6', 'Roof Type:', 'E6', 'External Floating'),
        ('D7', 'Floor Type:', 'E7', 'Cone Up (on ringwall)'),
        ('G4', 'Shell Material:', 'H4', 'A36'),
        ('G5', 'Floor Material:', 'H5', 'A36'),
        ('G6', 'Last Internal:', 'H6', '2021-08-15'),
        ('G7', 'Inspection Type:', 'H7', 'Internal (Out of Service)'),
    ]
    
    for label_cell, label, value_cell, value in tank_info:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, name='Arial', size=9)
        ws[value_cell] = value
        ws[value_cell].font = DATA_FONT
    
    # Inspection info
    ws['A9'] = f'Inspector: K. Thompson    Date: 2026-01-15    Instrument: Olympus 38DL+ S/N 34567'
    ws['A9'].font = DATA_FONT
    
    # Shell readings
    ws['A11'] = 'SHELL THICKNESS READINGS (mm)'
    ws['A11'].font = Font(bold=True, name='Arial', size=11)
    
    shell_headers = ['Course', 'Original\n(mm)', 'N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW', 'Min', 't-min\n(mm)', 'Status']
    
    for col, header in enumerate(shell_headers, 1):
        cell = ws.cell(row=12, column=col, value=header)
        apply_header_style(cell)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[12].height = 30
    
    # Tank shell data - typical for 40-year old crude tank
    shell_data = [
        [1, 25.40, 22.15, 22.08, 21.95, 22.12, 22.18, 21.88, 22.05, 22.11, 21.88, 12.70, 'Good'],
        [2, 22.23, 19.85, 19.92, 19.78, 19.88, 19.95, 19.72, 19.82, 19.91, 19.72, 11.12, 'Good'],
        [3, 19.05, 17.25, 17.18, 17.08, 17.22, 17.28, 17.02, 17.15, 17.21, 17.02, 9.53, 'Good'],
        [4, 15.88, 14.52, 14.48, 14.38, 14.55, 14.58, 14.32, 14.45, 14.51, 14.32, 7.94, 'Good'],
        [5, 12.70, 11.85, 11.78, 11.68, 11.82, 11.88, 11.62, 11.75, 11.81, 11.62, 6.35, 'Good'],
        [6, 9.53, 8.92, 8.85, 8.75, 8.88, 8.95, 8.68, 8.82, 8.88, 8.68, 4.76, 'Good'],
        [7, 9.53, 8.88, 8.82, 8.72, 8.85, 8.91, 8.65, 8.78, 8.85, 8.65, 4.76, 'Good'],
        [8, 9.53, 8.95, 8.88, 8.78, 8.92, 8.98, 8.72, 8.85, 8.92, 8.72, 4.76, 'Good'],
    ]
    
    for row_idx, row_data in enumerate(shell_data, 13):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell)
    
    # Floor readings
    ws['A23'] = 'FLOOR THICKNESS READINGS (mm) - See attached sketch for grid'
    ws['A23'].font = Font(bold=True, name='Arial', size=11)
    
    floor_headers = ['Grid', '1', '2', '3', '4', '5', '6', '7', '8', 'Row Min']
    for col, header in enumerate(floor_headers, 1):
        cell = ws.cell(row=24, column=col, value=header)
        apply_subheader_style(cell)
    
    # Floor data - Original 6.35mm floor, typical pitting
    floor_data = [
        ['A', 5.82, 5.78, 5.85, 5.75, 5.88, 5.72, 5.81, 5.79, 5.72],
        ['B', 5.75, 5.68, 5.72, 5.65, 5.78, 5.62, 5.71, 5.69, 5.62],
        ['C', 5.88, 5.82, 5.78, 5.71, 5.85, 5.68, 5.82, 5.78, 5.68],
        ['D', 5.72, 5.65, 5.58, 4.85, 5.68, 5.52, 5.65, 5.62, 4.85],  # Low reading
        ['E', 5.78, 5.72, 5.68, 5.62, 5.75, 5.58, 5.72, 5.68, 5.58],
        ['F', 5.85, 5.78, 5.75, 5.68, 5.82, 5.65, 5.78, 5.75, 5.65],
        ['G', 5.92, 5.85, 5.82, 5.75, 5.88, 5.72, 5.85, 5.82, 5.72],
        ['H', 5.88, 5.82, 5.78, 5.72, 5.85, 5.68, 5.81, 5.78, 5.68],
    ]
    
    for row_idx, row_data in enumerate(floor_data, 25):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell)
            # Highlight the low reading
            if value == 4.85:
                cell.fill = WARNING_FILL
    
    # Annular readings
    ws['A35'] = 'ANNULAR PLATE READINGS (mm) - Original: 12.70mm, t-min: 6.35mm'
    ws['A35'].font = Font(bold=True, name='Arial', size=11)
    
    annular_headers = ['Location', 'N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW']
    for col, header in enumerate(annular_headers, 1):
        cell = ws.cell(row=36, column=col, value=header)
        apply_subheader_style(cell)
    
    annular_data = [
        ['Inner Edge', 11.52, 11.48, 11.42, 11.55, 11.58, 11.38, 11.45, 11.51],
        ['Mid', 11.68, 11.62, 11.55, 11.71, 11.75, 11.52, 11.61, 11.68],
        ['Outer Edge', 11.85, 11.78, 11.72, 11.88, 11.92, 11.68, 11.78, 11.85],
    ]
    
    for row_idx, row_data in enumerate(annular_data, 37):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_data_style(cell)
    
    # Notes
    ws['A42'] = 'NOTES:'
    ws['A42'].font = Font(bold=True, name='Arial', size=10)
    ws['A43'] = '- Floor reading D4 (4.85mm) indicates localized pitting. Mark for monitoring or repair.'
    ws['A44'] = '- Original floor thickness: 6.35mm, t-min per API 653: 2.54mm (0.1")'
    ws['A45'] = '- All shell courses within acceptable limits. Next shell inspection: 10 years.'
    
    widths = {'A': 12, 'B': 10, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8, 'I': 8, 'J': 8, 'K': 10, 'L': 10, 'M': 10}
    set_column_widths(ws, widths)
    
    return ws


# ============================================================================
# ALTERNATIVE FORMATS (Different company styles)
# ============================================================================

def create_compact_piping_format(wb):
    """Compact single-page format used by some contractors"""
    ws = wb.create_sheet("Piping_Compact_Style")
    
    ws['A1'] = 'PIPING UT THICKNESS DATA SHEET'
    ws['A1'].font = Font(bold=True, name='Arial', size=12)
    
    # All header info in 2 rows
    compact_header = [
        'Circuit', 'Service', 'Material', 'Inspector', 'Date', 'Instrument S/N'
    ]
    compact_values = ['', '', '', '', '', '']
    
    for col, header in enumerate(compact_header, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_subheader_style(cell)
        ws.cell(row=4, column=col, value='')
        apply_data_style(ws.cell(row=4, column=col))
    
    # Very compact data table
    headers = ['CML', 'Comp', 'NPS', 'Nom', 'R1', 'R2', 'R3', 'R4', 'Min', 'Prev', 'Rate', 'Life', 'Note']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        apply_header_style(cell)
    
    for row in range(7, 32):
        for col in range(1, 14):
            apply_data_style(ws.cell(row=row, column=col))
    
    widths = {'A': 6, 'B': 10, 'C': 5, 'D': 6, 'E': 6, 'F': 6, 'G': 6, 'H': 6, 
              'I': 6, 'J': 6, 'K': 6, 'L': 6, 'M': 20}
    set_column_widths(ws, widths)
    
    return ws


def create_multipage_vessel_format(wb):
    """Detailed multi-zone vessel format"""
    ws = wb.create_sheet("Vessel_Detailed_Zones")
    
    ws['A1'] = 'PRESSURE VESSEL INSPECTION - DETAILED ZONE REPORT'
    ws['A1'].font = Font(bold=True, name='Arial', size=12)
    
    # This format has separate sections for each zone
    zones = [
        ('TOP HEAD', 5, ['TH-C', 'TH-N', 'TH-E', 'TH-S', 'TH-W']),
        ('SHELL COURSE 1', 12, ['SC1-N', 'SC1-NE', 'SC1-E', 'SC1-SE', 'SC1-S', 'SC1-SW', 'SC1-W', 'SC1-NW']),
        ('SHELL COURSE 2', 22, ['SC2-N', 'SC2-NE', 'SC2-E', 'SC2-SE', 'SC2-S', 'SC2-SW', 'SC2-W', 'SC2-NW']),
    ]
    
    for zone_name, start_row, cmls in zones:
        ws.cell(row=start_row, column=1, value=zone_name)
        ws.cell(row=start_row, column=1).font = Font(bold=True, name='Arial', size=10)
        
        # Sub-headers
        sub_headers = ['CML', 'Nominal', 't-min', 'Reading', 'Prev', 'Rate', 'Life', 'Condition']
        for col, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            apply_subheader_style(cell)
        
        for row_offset, cml in enumerate(cmls, 2):
            ws.cell(row=start_row + row_offset, column=1, value=cml)
            for col in range(1, 9):
                apply_data_style(ws.cell(row=start_row + row_offset, column=col))
    
    return ws


# ============================================================================
# MAIN GENERATION
# ============================================================================

def main():
    # Create workbook for piping
    wb_piping = Workbook()
    wb_piping.remove(wb_piping.active)  # Remove default sheet
    
    create_piping_template_style1(wb_piping)
    create_piping_template_style2(wb_piping)
    create_piping_sample_data(wb_piping)
    create_piping_sample2(wb_piping)
    create_compact_piping_format(wb_piping)
    
    wb_piping.save('/home/claude/cml_templates/Piping_CML_Templates.xlsx')
    print("Created: Piping_CML_Templates.xlsx")
    
    # Create workbook for vessels
    wb_vessel = Workbook()
    wb_vessel.remove(wb_vessel.active)
    
    create_vessel_template(wb_vessel)
    create_vessel_sample(wb_vessel)
    create_multipage_vessel_format(wb_vessel)
    
    wb_vessel.save('/home/claude/cml_templates/Vessel_CML_Templates.xlsx')
    print("Created: Vessel_CML_Templates.xlsx")
    
    # Create workbook for tanks
    wb_tank = Workbook()
    wb_tank.remove(wb_tank.active)
    
    create_tank_template(wb_tank)
    create_tank_sample(wb_tank)
    
    wb_tank.save('/home/claude/cml_templates/Tank_CML_Templates.xlsx')
    print("Created: Tank_CML_Templates.xlsx")
    
    print("\nAll CML templates generated successfully!")


if __name__ == '__main__':
    main()

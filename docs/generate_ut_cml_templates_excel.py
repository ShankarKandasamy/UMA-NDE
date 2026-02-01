import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
TITLE_FILL = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row_idx: int, bold: bool = True, center: bool = True) -> None:
    """
    Apply a consistent header style to the given row.
    """
    for cell in ws[row_idx]:
        cell.font = Font(bold=bold, color="000000")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_columns(ws, min_width: int = 8, max_width: int = 40) -> None:
    """
    Rough auto-fit for column widths based on cell contents.
    """
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except Exception:
                value = ""
            max_length = max(max_length, len(value))
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width


def _add_title(ws, title: str, row_idx: int = 1, col_span: int = 8) -> None:
    """
    Add a merged, styled title at the top of the sheet.
    """
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_span)
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = TITLE_FILL
    cell.border = THIN_BORDER


def sheet_1_piping_circuit_thickness(wb: Workbook) -> None:
    ws = wb.create_sheet(title="01_PipingCircuit")

    _add_title(ws, "1. Piping Circuit Thickness Sheet (API 570 / 574)", col_span=8)

    # Header block
    ws["A3"] = "Line No"
    ws["B3"] = ""
    ws["C3"] = "Circuit ID"
    ws["D3"] = ""
    ws.merge_cells("A3:B3")
    ws.merge_cells("C3:D3")

    ws["A4"] = "Service"
    ws["B4"] = ""
    ws["C4"] = "Size"
    ws["D4"] = ""
    ws["E4"] = "Sch"
    ws["F4"] = ""
    ws["G4"] = "Material"
    ws["H4"] = ""
    ws.merge_cells("A4:B4")
    ws.merge_cells("C4:D4")
    ws.merge_cells("E4:F4")
    ws.merge_cells("G4:H4")

    for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=8):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["C3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)
    ws["C4"].font = Font(bold=True)
    ws["E4"].font = Font(bold=True)
    ws["G4"].font = Font(bold=True)

    # Table
    headers = ["CML", "Location Description", "Nom t", "Meas t", "Tmin", "CR", "RL", "Notes"]
    ws.append([])
    ws.append(headers)  # row 6
    _style_header_row(ws, 6)

    # Add a few empty data rows for convenience
    for _ in range(20):
        ws.append([""] * len(headers))

    # Borders for data area
    for row in ws.iter_rows(min_row=7, max_row=26, min_col=1, max_col=8):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    _autofit_columns(ws)


def sheet_2_orientation_based_grid(wb: Workbook) -> None:
    ws = wb.create_sheet(title="02_OrientationGrid")

    _add_title(ws, "2. Orientation-Based CML Grid", col_span=4)

    ws["A3"] = "Line No"
    ws["B3"] = ""
    ws["C3"] = "CML ID"
    ws["D3"] = ""
    ws.merge_cells("A3:B3")
    ws.merge_cells("C3:D3")

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["C3"].font = Font(bold=True)

    ws["A5"] = "12 o'clock"
    ws["B5"] = ""
    ws["A6"] = "3 o'clock"
    ws["B6"] = ""
    ws["A7"] = "6 o'clock"
    ws["B7"] = ""
    ws["A8"] = "9 o'clock"
    ws["B8"] = ""

    for row in ws.iter_rows(min_row=5, max_row=8, min_col=1, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in range(5, 9):
        ws[f"A{r}"].font = Font(bold=True)

    ws["A10"] = "Minimum Thickness"
    ws["B10"] = ""
    ws.merge_cells("A10:B10")
    ws["A10"].font = Font(bold=True)
    ws["A10"].border = THIN_BORDER
    ws["B10"].border = THIN_BORDER
    ws["A10"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A12"] = "Comments"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = ""
    ws["A14"] = ""
    ws["A15"] = ""
    ws.merge_cells("A13:D15")
    for row in ws.iter_rows(min_row=12, max_row=15, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autofit_columns(ws)


def sheet_3_pipe_shoe_cui(wb: Workbook) -> None:
    ws = wb.create_sheet(title="03_PipeShoe_CUI")

    _add_title(ws, "3. Pipe Shoe / CUI Thickness Sheet", col_span=6)

    ws["A3"] = "Line No"
    ws["B3"] = ""
    ws["C3"] = "Pipe Shoe ID"
    ws["D3"] = ""
    ws.merge_cells("A3:B3")
    ws.merge_cells("C3:D3")

    ws["A4"] = "Support Type"
    ws["B4"] = ""
    ws["C4"] = "Insulation Removed (Y/N)"
    ws["D4"] = ""
    ws.merge_cells("A4:B4")
    ws.merge_cells("C4:D4")

    for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["C3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)
    ws["C4"].font = Font(bold=True)

    headers = ["Top", "Bottom", "Side A", "Side B", "Min t"]
    ws.append([])
    ws.append(headers)  # row 6
    _style_header_row(ws, 6)

    for _ in range(10):
        ws.append([""] * len(headers))

    for row in ws.iter_rows(min_row=7, max_row=16, min_col=1, max_col=5):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["A18"] = "Observations"
    ws["A18"].font = Font(bold=True)
    ws.merge_cells("A19:F23")
    for row in ws.iter_rows(min_row=18, max_row=23, min_col=1, max_col=6):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autofit_columns(ws)


def sheet_4_isometric_legend(wb: Workbook) -> None:
    ws = wb.create_sheet(title="04_Isometric_Legend")

    _add_title(ws, "4. Isometric + CML Legend", col_span=4)

    ws["A3"] = "(SKETCH AREA – Inspector draws piping run & marks CMLs)"
    ws.merge_cells("A3:D15")
    for row in ws.iter_rows(min_row=3, max_row=15, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws["A17"] = "Legend"
    ws["A17"].font = Font(bold=True)

    headers = ["CML", "Thickness", "Comments"]
    ws.append([])
    ws.append(headers)  # row 19
    _style_header_row(ws, 19)
    for _ in range(30):
        ws.append([""] * len(headers))
    for row in ws.iter_rows(min_row=20, max_row=49, min_col=1, max_col=3):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _autofit_columns(ws)


def sheet_5_shutdown_field(wb: Workbook) -> None:
    ws = wb.create_sheet(title="05_Shutdown_Field")

    _add_title(ws, "5. Shutdown Field Sheet (Minimal)", col_span=7)

    ws["A3"] = "Job #"
    ws["B3"] = ""
    ws["C3"] = "Date"
    ws["D3"] = datetime.date.today()
    ws.merge_cells("A3:B3")
    ws.merge_cells("C3:D3")

    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["C3"].font = Font(bold=True)

    headers = ["CML", "t1", "t2", "t3", "t4", "Min t", "Comments"]
    ws.append([])
    ws.append(headers)  # row 5
    _style_header_row(ws, 5)

    for _ in range(40):
        ws.append([""] * len(headers))
    for row in ws.iter_rows(min_row=6, max_row=45, min_col=1, max_col=7):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column < 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _autofit_columns(ws)


def sheet_6_owner_mandated(wb: Workbook) -> None:
    ws = wb.create_sheet(title="06_Owner_Corrosion")

    _add_title(ws, "6. Owner-Mandated Corrosion Sheet", col_span=7)

    ws["A3"] = "Client"
    ws["B3"] = ""
    ws["C3"] = "Unit"
    ws["D3"] = ""
    ws.merge_cells("A3:B3")
    ws.merge_cells("C3:D3")

    ws["A4"] = "Risk Class"
    ws["B4"] = "High / Medium / Low"
    ws.merge_cells("A4:B4")

    for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["C3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)

    headers = ["CML", "Design t", "Last t", "Curr t", "Δt", "CR", "Next Insp"]
    ws.append([])
    ws.append(headers)  # row 6
    _style_header_row(ws, 6)
    for _ in range(40):
        ws.append([""] * len(headers))
    for row in ws.iter_rows(min_row=7, max_row=46, min_col=1, max_col=7):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _autofit_columns(ws)


def sheet_7_vessel_shell_grid(wb: Workbook) -> None:
    ws = wb.create_sheet(title="07_PV_ShellGrid")

    _add_title(ws, "7. Pressure Vessel – Shell Course Grid (API 510)", col_span=5)

    ws["A3"] = "Vessel ID"
    ws["B3"] = ""
    ws.merge_cells("A3:B3")
    for cell in ws["A3:B3"]:
        for c in cell:
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)

    # Course 1
    ws["A5"] = "Shell Course 1"
    ws["A5"].font = Font(bold=True)
    headers = ["N", "E", "S", "W"]
    ws.append([])
    ws.append(headers)  # row 7
    _style_header_row(ws, 7)
    ws.append(["", "", "", ""])  # row 8
    for row in ws.iter_rows(min_row=8, max_row=8, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Course 2
    ws["A10"] = "Shell Course 2"
    ws["A10"].font = Font(bold=True)
    ws.append([])
    ws.append(headers)  # row 12
    _style_header_row(ws, 12)
    ws.append(["", "", "", ""])  # row 13
    for row in ws.iter_rows(min_row=13, max_row=13, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["A15"] = "Minimum Required Thickness"
    ws["B15"] = ""
    ws.merge_cells("A15:B15")
    for cell in ws["A15:B15"]:
        for c in cell:
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws["A15"].font = Font(bold=True)

    _autofit_columns(ws)


def sheet_8_vessel_nozzles(wb: Workbook) -> None:
    ws = wb.create_sheet(title="08_PV_Nozzles")

    _add_title(ws, "8. Pressure Vessel – Nozzle Thickness Sheet", col_span=6)

    ws["A3"] = "Vessel ID"
    ws["B3"] = ""
    ws.merge_cells("A3:B3")
    for cell in ws["A3:B3"]:
        for c in cell:
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)

    headers = ["Nozzle", "Orientation", "Nom t", "Meas t", "Tmin", "Remarks"]
    ws.append([])
    ws.append(headers)  # row 5
    _style_header_row(ws, 5)
    for _ in range(40):
        ws.append([""] * len(headers))
    for row in ws.iter_rows(min_row=6, max_row=45, min_col=1, max_col=6):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column < 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _autofit_columns(ws)


def sheet_9_pv_engineering_eval(wb: Workbook) -> None:
    ws = wb.create_sheet(title="09_PV_EngEval")

    _add_title(ws, "9. Pressure Vessel – Engineering Evaluation Hybrid", col_span=4)

    ws["A3"] = "Location"
    ws["B3"] = ""
    ws.merge_cells("A3:B3")
    ws["A4"] = "Measured t"
    ws["B4"] = ""
    ws.merge_cells("A4:B4")
    ws["A5"] = "Tmin"
    ws["B5"] = ""
    ws.merge_cells("A5:B5")
    for row in ws.iter_rows(min_row=3, max_row=5, min_col=1, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)
    ws["A5"].font = Font(bold=True)

    ws["A7"] = "Engineering Notes"
    ws["A7"].font = Font(bold=True)
    ws.merge_cells("A8:D16")
    for row in ws.iter_rows(min_row=7, max_row=16, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws["A18"] = "Acceptability per API 510"
    ws["A18"].font = Font(bold=True)
    ws["A19"] = "[ ] ACCEPTABLE"
    ws["B19"] = "[ ] NOT ACCEPTABLE"
    for row in ws.iter_rows(min_row=18, max_row=19, min_col=1, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    _autofit_columns(ws)


def sheet_10_tank_shell_compass(wb: Workbook) -> None:
    ws = wb.create_sheet(title="10_Tank_ShellCompass")

    _add_title(ws, "10. Storage Tank – Shell Compass Table (API 653)", col_span=4)

    ws["A3"] = "Tank ID"
    ws["B3"] = ""
    ws["C3"] = "Shell Course"
    ws["D3"] = ""
    ws.merge_cells("A3:B3")
    ws.merge_cells("C3:D3")
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["A3"].font = Font(bold=True)
    ws["C3"].font = Font(bold=True)

    ws["A5"] = "0° (N)"
    ws["B5"] = ""
    ws["A6"] = "90° (E)"
    ws["B6"] = ""
    ws["A7"] = "180° (S)"
    ws["B7"] = ""
    ws["A8"] = "270° (W)"
    ws["B8"] = ""
    for row in ws.iter_rows(min_row=5, max_row=8, min_col=1, max_col=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
    for r in range(5, 9):
        ws[f"A{r}"].font = Font(bold=True)

    ws["A10"] = "Minimum Thickness"
    ws["B10"] = ""
    ws.merge_cells("A10:B10")
    for cell in ws["A10:B10"]:
        for c in cell:
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws["A10"].font = Font(bold=True)

    _autofit_columns(ws)


def sheet_11_tank_multi_course(wb: Workbook) -> None:
    ws = wb.create_sheet(title="11_Tank_MultiCourse")

    _add_title(ws, "11. Storage Tank – Multi-Course Summary", col_span=5)

    headers = ["Course", "Nom t", "Min t", "Avg t", "% Loss"]
    ws.append([])
    ws.append(headers)  # row 3
    _style_header_row(ws, 3)
    for _ in range(20):
        ws.append([""] * len(headers))
    for row in ws.iter_rows(min_row=4, max_row=23, min_col=1, max_col=5):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    _autofit_columns(ws)


def sheet_12_tank_inspector_notebook(wb: Workbook) -> None:
    ws = wb.create_sheet(title="12_Tank_InspectorNotes")

    _add_title(ws, "12. Storage Tank – Inspector Notebook Page", col_span=4)

    ws["A3"] = (
        "Free-form page for:\n"
        "• Handwritten notes\n"
        "• Sketches\n"
        "• Arrows\n"
        "• Circled values\n"
        "• “Bad area here” annotations"
    )
    ws.merge_cells("A3:D30")
    for row in ws.iter_rows(min_row=3, max_row=30, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _autofit_columns(ws)


def build_workbook() -> Workbook:
    wb = Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    sheet_1_piping_circuit_thickness(wb)
    sheet_2_orientation_based_grid(wb)
    sheet_3_pipe_shoe_cui(wb)
    sheet_4_isometric_legend(wb)
    sheet_5_shutdown_field(wb)
    sheet_6_owner_mandated(wb)
    sheet_7_vessel_shell_grid(wb)
    sheet_8_vessel_nozzles(wb)
    sheet_9_pv_engineering_eval(wb)
    sheet_10_tank_shell_compass(wb)
    sheet_11_tank_multi_course(wb)
    sheet_12_tank_inspector_notebook(wb)

    return wb


def main(output_path: str = "UT_CML_Templates.xlsx") -> None:
    wb = build_workbook()
    wb.save(output_path)
    print(f"Saved UT CML templates workbook to: {output_path}")


if __name__ == "__main__":
    main()

